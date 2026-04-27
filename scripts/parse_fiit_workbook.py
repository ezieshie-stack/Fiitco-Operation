#!/usr/bin/env python3

import json
import re
import sys
from collections import OrderedDict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


DAY_LABELS = {
    "Monday": "Mon",
    "Tuesday": "Tue",
    "Wednesday": "Wed",
    "Thursday": "Thu",
    "Friday": "Fri",
    "Saturday": "Sat",
    "Sunday": "Sun",
}

SCHEDULE_ALIAS_TO_CLASS_ID = {
    "Boxing Basics": "CLS-04",
    "Lunch Box": "CLS-04",
    "Bag Work": "CLS-04",
    "Weekend Warriors": "CLS-04",
    "Advanced Boxing": "CLS-05",
    "Fight Night Prep": "CLS-05",
    "Open Sparring": "CLS-05",
    "Hybrid Crusher": "CLS-06",
    "Hybrid & Pilates": "CLS-06",
    "Yoga & Box": "CLS-06",
    "HIT Circuit": "CLS-06",
    "Late Night HIT": "CLS-06",
    "Sunday Sweat": "CLS-06",
    "Pilates Box": "CLS-07",
    "Yoga Circuit": "CLS-08",
    "Power Hour": "CLS-01",
    "Strength & Conditioning": "CLS-01",
    "Full Body Strength": "CLS-03",
}


def fail(message: str) -> None:
    raise SystemExit(message)


def find_sheet(workbook, fragment: str):
    fragment_lower = fragment.lower()
    for sheet_name in workbook.sheetnames:
        if fragment_lower in sheet_name.lower():
            return workbook[sheet_name]
    fail(f"Could not find worksheet containing '{fragment}'")


def clean_string(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return str(value)


def split_csv(value):
    text = clean_string(value)
    if not text or text.lower() == "none":
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def yn_to_bool(value):
    text = clean_string(value)
    if text is None:
        return False
    return text.upper() == "Y"


def normalize_time(value):
    text = clean_string(value)
    if text is None:
        return None
    return datetime.strptime(text.upper(), "%I:%M %p").strftime("%H:%M")


def add_minutes(time_value: str, minutes: int) -> str:
    hour, minute = [int(part) for part in time_value.split(":")]
    total = hour * 60 + minute + minutes
    return f"{total // 60:02d}:{total % 60:02d}"


def iso_date(value) -> str:
    text = clean_string(value)
    if text is None:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    return datetime.strptime(text, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")


def iso_datetime_from_text(value: str):
    text = clean_string(value)
    if not text:
        return None
    return datetime.strptime(text, "%B %d, %Y").strftime("%Y-%m-%dT00:00:00.000Z")


def parse_class_name(raw_name: str) -> str:
    return re.sub(r"^[^\w]+", "", raw_name).strip()


def resolve_class(raw_name: str, classes_by_id, class_id_by_name):
    class_name = parse_class_name(raw_name)
    class_id = class_id_by_name.get(class_name) or SCHEDULE_ALIAS_TO_CLASS_ID.get(class_name)
    if not class_id:
        fail(f"Could not resolve class name '{class_name}'")
    class_row = classes_by_id.get(class_id)
    if not class_row:
        fail(f"Resolved class '{class_name}' to '{class_id}', but that class is missing from the library")
    return class_id, class_name, class_row


def extract_rows(sheet, start_row: int):
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if any(cell is not None for cell in row):
            yield row


def parse_categories(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        rows.append({
            "categoryId": clean_string(row[0]) or "",
            "name": clean_string(row[1]) or "",
            "colorCode": clean_string(row[2]) or "",
            "emoji": clean_string(row[3]) or "",
            "description": clean_string(row[4]) or "",
            "active": yn_to_bool(row[5]),
        })
    return rows


def parse_subcategories(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        rows.append({
            "subcategoryId": clean_string(row[0]) or "",
            "categoryId": clean_string(row[1]) or "",
            "categoryName": clean_string(row[2]) or "",
            "name": clean_string(row[3]) or "",
            "description": clean_string(row[4]) or "",
            "active": yn_to_bool(row[5]),
        })
    return rows


def parse_classes(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        rows.append({
            "classId": clean_string(row[0]) or "",
            "categoryId": clean_string(row[1]) or "",
            "categoryName": clean_string(row[2]) or "",
            "subcategoryName": clean_string(row[3]),
            "name": clean_string(row[4]) or "",
            "tier": clean_string(row[5]) or "",
            "durationMinutes": int(row[6] or 0),
            "description": clean_string(row[7]) or "",
            "active": yn_to_bool(row[9]),
        })
    return rows


def parse_tiers(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        rows.append({
            "tierId": clean_string(row[0]) or "",
            "name": clean_string(row[1]) or "",
            "description": clean_string(row[2]) or "",
            "recommendedFor": clean_string(row[3]) or "",
            "colorCode": clean_string(row[4]) or "",
        })
    return rows


def parse_equipment(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        rows.append({
            "equipmentId": clean_string(row[0]) or "",
            "name": clean_string(row[1]) or "",
            "category": clean_string(row[2]) or "",
            "quantityAvailable": int(row[3] or 0),
            "location": clean_string(row[4]) or "",
            "notes": clean_string(row[5]),
            "active": yn_to_bool(row[6]),
        })
    return rows


def parse_pathways(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        rows.append({
            "pathwayId": clean_string(row[0]) or "",
            "title": clean_string(row[1]) or "",
            "category": clean_string(row[2]) or "",
            "targetTier": clean_string(row[3]) or "",
            "durationWeeks": int(str(row[4]).strip()),
            "goal": clean_string(row[5]) or "",
            "description": clean_string(row[6]) or "",
            "active": yn_to_bool(row[7]),
        })
    return rows


def parse_instructors(master_sheet, roster_sheet):
    master_by_name = {}
    for row in extract_rows(master_sheet, 3):
        full_name = clean_string(row[1])
        if not full_name:
            continue
        master_by_name[full_name] = {
            "displayName": clean_string(row[2]),
            "status": clean_string(row[7]) or "Active",
            "joinDate": clean_string(row[8]) or "",
            "notes": clean_string(row[9]),
        }

    rows = []
    for row in extract_rows(roster_sheet, 3):
        full_name = clean_string(row[1])
        master = master_by_name.get(full_name, {})
        display_name = master.get("displayName") or (full_name.split(" ")[0] if full_name else "")
        rows.append({
            "instructorId": clean_string(row[0]) or "",
            "fullName": full_name or "",
            "displayName": display_name,
            "specialisations": split_csv(row[2]),
            "certifications": split_csv(row[6]),
            "email": clean_string(row[5]) or "",
            "phone": clean_string(row[4]) or "",
            "status": master.get("status") or "Active",
            "joinDate": master.get("joinDate") or "",
            "notes": master.get("notes"),
        })
    return rows


def parse_availability(sheet, instructors_by_name):
    rows = []
    current_instructor = None
    header_row = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
    day_headers = [clean_string(value) for value in header_row][2:9]

    for row in extract_rows(sheet, 4):
        label = clean_string(row[1])
        if label and all(cell is None for cell in row[2:9]):
            current_instructor = label
            if current_instructor not in instructors_by_name:
                fail(f"Availability references unknown instructor '{current_instructor}'")
            continue

        if not current_instructor or not label or label.startswith("Instructor /"):
            continue

        if re.match(r"^\d{1,2}:\d{2} [AP]M$", label):
            start_time = normalize_time(label)
            end_time = add_minutes(start_time, 60)
            instructor = instructors_by_name[current_instructor]

            for day_name, cell_value in zip(day_headers, row[2:9]):
                value = clean_string(cell_value)
                if value not in {"✅", "❌"}:
                    continue
                rows.append({
                    "instructorId": instructor["instructorId"],
                    "instructorName": instructor["fullName"],
                    "dayOfWeek": DAY_LABELS[day_name],
                    "startTime": start_time,
                    "endTime": end_time,
                    "available": value == "✅",
                    "notes": None,
                })
    return rows


def parse_week_anchor(sheet):
    title = clean_string(sheet["A1"].value) or ""
    match = re.search(r"Week of ([A-Za-z]+ \d{1,2}, \d{4})", title)
    if not match:
        fail("Could not parse weekly schedule anchor date")
    anchor = datetime.strptime(match.group(1), "%B %d, %Y")
    return anchor - timedelta(days=anchor.weekday())


def parse_weekly_schedule(sheet, classes_by_id, class_id_by_name, instructors_by_name):
    rows = []
    monday = parse_week_anchor(sheet)
    header_row = next(sheet.iter_rows(min_row=4, max_row=4, values_only=True))
    day_headers = [clean_string(value) for value in header_row][1:8]

    for row in extract_rows(sheet, 5):
        first_cell = clean_string(row[0])
        if first_cell is None or first_cell.startswith("📊 Weekly Total") or first_cell.startswith("⚠️ 10-Minute"):
            break
        if not re.match(r"^\d{1,2}:\d{2} [AP]M$", first_cell):
            continue

        start_time = normalize_time(first_cell)
        for offset, (day_name, cell_value) in enumerate(zip(day_headers, row[1:8])):
            cell_text = clean_string(cell_value)
            if not cell_text or cell_text == "—":
                continue

            lines = [line.strip() for line in str(cell_text).splitlines() if line.strip()]
            if len(lines) < 2:
                fail(f"Could not parse weekly schedule cell '{cell_text}'")

            _, class_name, class_row = resolve_class(lines[0], classes_by_id, class_id_by_name)
            instructor_match = re.match(r"^(.*?)\s*\[(\d+)\]$", lines[1])
            if not instructor_match:
                fail(f"Could not parse instructor/capacity line '{lines[1]}'")

            instructor_name = instructor_match.group(1).strip()
            instructor = instructors_by_name.get(instructor_name)
            if not instructor:
                fail(f"Weekly schedule references unknown instructor '{instructor_name}'")

            rows.append({
                "date": (monday + timedelta(days=offset)).strftime("%Y-%m-%d"),
                "dayOfWeek": DAY_LABELS[day_name],
                "startTime": start_time,
                "endTime": add_minutes(start_time, int(class_row["durationMinutes"])),
                "classId": class_row["classId"],
                "className": class_name,
                "categoryName": class_row["categoryName"],
                "instructorId": instructor["instructorId"],
                "instructorName": instructor["displayName"],
                "capacity": int(instructor_match.group(2)),
                "status": "Scheduled",
                "bufferViolation": False,
                "bufferOverrideAcknowledged": False,
            })
    return rows, monday


def parse_class_programs(sheet, week_of, classes_by_id, class_id_by_name, instructors_by_name):
    rows = []
    current = None
    current_blocks = []

    def flush_current(footer_text=None):
        nonlocal current, current_blocks
        if not current:
            return
        submitted_at = None
        if footer_text:
            created_match = re.search(r"Created:\s*([A-Za-z]+ \d{1,2}, \d{4})", footer_text)
            if created_match:
                submitted_at = iso_datetime_from_text(created_match.group(1))
        rows.append({
            "classId": current["classId"],
            "className": current["className"],
            "instructorId": current["instructorId"],
            "instructorName": current["instructorName"],
            "weekOf": week_of,
            "blocks": current_blocks,
            "status": "Submitted",
            "submittedAt": submitted_at,
            "approvedAt": None,
            "approvedBy": None,
            "notes": None,
        })
        current = None
        current_blocks = []

    for row in extract_rows(sheet, 1):
        first_cell = clean_string(row[0])
        if first_cell and first_cell.startswith("CLASS: "):
            flush_current()
            match = re.match(r"^CLASS:\s*(.*?)\s*[—-]\s*(.*?)\s*[—-]\s*Instructor:\s*(.*)$", first_cell)
            if not match:
                fail(f"Could not parse class program header '{first_cell}'")
            _, class_name, class_row = resolve_class(match.group(1), classes_by_id, class_id_by_name)
            instructor_name = match.group(3).strip()
            instructor = instructors_by_name.get(instructor_name)
            if not instructor:
                fail(f"Class program references unknown instructor '{instructor_name}'")
            current = {
                "classId": class_row["classId"],
                "className": class_name,
                "instructorId": instructor["instructorId"],
                "instructorName": instructor["fullName"],
            }
            continue

        if not current or first_cell == "Class Name":
            continue

        if first_cell and first_cell.startswith("Total class duration:"):
            flush_current(first_cell)
            continue

        phase = clean_string(row[3])
        title = clean_string(row[4])
        if not phase or not title:
            continue

        instructions = clean_string(row[6]) or ""
        note = clean_string(row[8])
        if note:
            instructions = f"{instructions}\nNote: {note}" if instructions else f"Note: {note}"

        duration_match = re.search(r"(\d+)", clean_string(row[5]) or "")
        current_blocks.append({
            "blockType": phase,
            "exerciseName": None,
            "durationMinutes": int(duration_match.group(1)) if duration_match else 0,
            "description": title,
            "equipment": split_csv(row[7]),
            "instructions": instructions,
        })

    flush_current()
    return rows


def parse_delivery_logs(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        if clean_string(row[0]) == "DELIVERY LOG SUMMARY":
            break

        member_feedback = clean_string(row[13])
        logged_by = clean_string(row[14])
        note_parts = []
        if member_feedback:
            note_parts.append(f"Member Feedback: {member_feedback}")
        if logged_by:
            note_parts.append(f"Logged By: {logged_by}")
        extra_notes = clean_string(row[15])
        if extra_notes:
            note_parts.append(extra_notes)

        rows.append({
            "date": iso_date(row[1]) or "",
            "classId": clean_string(row[2]) or "",
            "className": clean_string(row[3]) or "",
            "categoryName": clean_string(row[4]) or "",
            "instructorId": clean_string(row[5]) or "",
            "instructorName": clean_string(row[6]) or "",
            "wasPlanned": yn_to_bool(row[7]),
            "actualAttendance": int(row[8] or 0),
            "maxCapacity": int(row[9] or 0),
            "programFollowed": yn_to_bool(row[11]),
            "variationsMade": clean_string(row[12]),
            "notes": "\n".join(note_parts) if note_parts else None,
        })
    return rows


def parse_client_journeys(sheet):
    grouped = OrderedDict()
    status_by_journey = {}

    for row in extract_rows(sheet, 3):
        journey_id = clean_string(row[0])
        if not journey_id:
            continue

        if journey_id not in grouped:
            grouped[journey_id] = {
                "journeyId": journey_id,
                "title": clean_string(row[1]) or "",
                "goalType": clean_string(row[2]) or "",
                "pathwayId": clean_string(row[3]) or "",
                "weeks": [],
            }
            status_by_journey[journey_id] = []

        grouped[journey_id]["weeks"].append({
            "weekNumber": int(str(row[4]).strip()),
            "classId": clean_string(row[5]) or "",
            "className": clean_string(row[6]) or "",
            "focus": clean_string(row[8]) or "",
            "notes": clean_string(row[9]),
        })
        status_by_journey[journey_id].append(clean_string(row[10]) or "")

    rows = []
    for journey_id, journey in grouped.items():
        statuses = status_by_journey[journey_id]
        journey["active"] = any(status not in {"Completed", "Archived", "Cancelled"} for status in statuses)
        rows.append(journey)
    return rows


def parse_exercises(sheet):
    rows = []
    for row in extract_rows(sheet, 3):
        number = int(row[0])
        rows.append({
            "exerciseId": f"EXC-{number:03d}",
            "name": clean_string(row[2]) or "",
            "category": clean_string(row[1]) or "",
            "subcategory": None,
            "tier": clean_string(row[5]),
            "description": clean_string(row[3]) or "",
            "equipment": split_csv(row[4]),
            "active": True,
        })
    return rows


def main():
    workbook_path = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) > 1 else None
    if not workbook_path or not workbook_path.exists():
        fail("Usage: parse_fiit_workbook.py /absolute/path/to/FIIT_Co_Class_Management_Tool_v4_FINAL.xlsx")

    workbook = load_workbook(workbook_path, data_only=True)

    categories = parse_categories(find_sheet(workbook, "Category Library"))
    subcategories = parse_subcategories(find_sheet(workbook, "Subcategory Library"))
    classes = parse_classes(find_sheet(workbook, "Class Library"))
    tiers = parse_tiers(find_sheet(workbook, "Tiers Library"))
    equipment = parse_equipment(find_sheet(workbook, "Equipment Library"))
    pathways = parse_pathways(find_sheet(workbook, "Pathway Library"))
    instructors = parse_instructors(
        find_sheet(workbook, "Instructor Library"),
        find_sheet(workbook, "Instructors"),
    )

    instructors_by_name = {row["fullName"]: row for row in instructors}
    classes_by_id = {row["classId"]: row for row in classes}
    class_id_by_name = {row["name"]: row["classId"] for row in classes}

    availability = parse_availability(find_sheet(workbook, "Availability & Subs"), instructors_by_name)
    weekly_schedule, monday = parse_weekly_schedule(
        find_sheet(workbook, "Weekly Schedule"),
        classes_by_id,
        class_id_by_name,
        instructors_by_name,
    )
    class_programs = parse_class_programs(
        find_sheet(workbook, "Class Program"),
        monday.strftime("%Y-%m-%d"),
        classes_by_id,
        class_id_by_name,
        instructors_by_name,
    )
    delivery_log = parse_delivery_logs(find_sheet(workbook, "Class Delivery Log"))
    client_journeys = parse_client_journeys(find_sheet(workbook, "Client Journey"))
    exercises = parse_exercises(find_sheet(workbook, "Exercise Library"))

    payload = {
        "categories": categories,
        "subcategories": subcategories,
        "classes": classes,
        "instructors": instructors,
        "tiers": tiers,
        "equipment": equipment,
        "pathways": pathways,
        "exercises": exercises,
        "weeklySchedule": weekly_schedule,
        "classPrograms": class_programs,
        "deliveryLog": delivery_log,
        "clientJourneys": client_journeys,
        "availability": availability,
    }

    json.dump(payload, sys.stdout, ensure_ascii=True)


if __name__ == "__main__":
    main()
